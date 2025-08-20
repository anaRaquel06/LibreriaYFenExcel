"""

@author: Ana López
"""

import yfinance as yf
from datetime import datetime, timedelta
from openpyxl import Workbook
import os

# Lista de acciones del codigo 
tickers = [
    "AA","AAL","AAPL","AMM.TO","ABBV","ABNB","ACTINVRB.MX","AC","AFRM",
    "AGNC","ALFAA.MX","ALPEKA.MX","ALSEA.MX","AMAT","AMD","AMX","AMZN",
    "APA","ASURB.MX","ATER","ATOS","AIY.DE","AVGO","AXP","BABA","BAC",
    "BA","BBAJIOO.MX","BIMBOA.MX","BMY","BNGO","CAT","CCL",
    "CEMEXCPO.MX","CHDRAUIB.MX","CLF","COST","CRM","CSCO",
    "CUERVO.MX","CVS","CVX","C","DAL","DIS","DVN","ELEKTRA.MX","ETSY",
    "FANG","FCX","FDX","FEMSAUBD.MX","FIBRAMQ12.MX","FIBRAPL14.MX",
    "FSLR","FUBO","FUNO11.MX","F","GAPB.MX","GCARSOA1.MX","GCC",
    "GENTERA.MX","GE","GFINBURO.MX","GFNORTEO.MX","GILD","GMEXICOB.MX",
    "GME","GM","GOLD","GOOGL","GRUMAB.MX","HD","INTC","JNJ","JPM",
    "KIMBERA.MX","KOFUBL.MX","KO","LABB.MX",
    "LASITEB-1.MX","LCID","LIVEPOLC-1.MX","LLY","LUV","LVS","LYFT","MARA",
    "MA","MCD","MEGACPO.MX","MELIN.MX","META","MFRISCOA-1.MX","MGM",
    "MRK","MRNA","MRO","MSFT","MU","NCLHN.MX","NFLX","NKE","NKLA","NUN.MX",
    "NVAX","NVDA","OMAB.MX","ORBIA.MX","ORCL","OXY1.MX","PARA","PBRN.MX","PE&OLES.MX",
    "PEP","PFE","PG","PINFRA.MX","PINS","PLTR","PYPL","QCOM","Q.MX","RCL",
    "RIOT","RIVN","ROKU","RA.MX","SBUX","SHOP","SITES1A-1.MX","SKLZ",
    "SOFI","SPCE","SQ","TALN.MX","TERRA13.MX","TGT","TLEVISACPO.MX","TMO",
    "TSLA","TSMN.MX","TWLO","TX","T","UAL","UBER","UNH","UPST","VESTA.MX",
    "VOLARA.MX","VZ","V","WALMEX.MX","WFC","WMT","WYNN","XOM","X","ZM"
]

# Fechas de los últimos 180 días esta vez 
end_date = datetime.today()
start_date = end_date - timedelta(days=180)

# Crear libro de Excel con Openpyxl y agregar el titulo 
wb = Workbook()
ws = wb.active
ws.title = "Resumen 180 días"
headers = ["Ticker", "Fecha", "Apertura", "Alta", "Baja", "Cierre", "Volumen"]
ws.append(headers)

# Descargar datos de yahoo finance con los tickers otra vez  
for ticker in tickers:
    try:
        stock = yf.Ticker(ticker)
        df = stock.history(start=start_date, end=end_date)

        if df.empty: 
            print(f"⚠️ Sin datos para {ticker}")
            continue

        df.reset_index(inplace=True)

        for _, row in df.iterrows():
            ws.append([
                ticker,
                row["Date"].strftime('%Y-%m-%d'),
                round(row["Open"], 2),
                round(row["High"], 2),
                round(row["Low"], 2),
                round(row["Close"], 2),
                int(row["Volume"])
            ])
        print(f"✅ {ticker} completado") #otros indicadores (se ve amigable) 🟢👍💯
    except Exception as e:
        print(f"❌ Error con {ticker}: {e}")  #otras opciones que podemos usar aqwui 🚨🚧⛔📢💣

# Tomamos todos los datos menos el encabezado
all_data = list(ws.iter_rows(min_row=2, values_only=True))

# Ordenamos:
#   1) Ticker (acciones juntas) aqui 
#   2) Fecha ascendente (dentro de cada ticker) ordenadas 
sorted_data = sorted(all_data, key=lambda row: (row[0], row[1]))

# Aqui se  borraron todas las filas menos el encabezado
ws.delete_rows(2, ws.max_row)

# Reescribir los datos ordenados
for row in sorted_data:
    ws.append(row)

# Guardar archivo
filename = "historico_ult_mes.xlsx" #le puse el mismo nombre de archivo para no reescribir tod pero son de los ultimos 180 
wb.save(filename)

print(f"\n📁 Archivo guardado como: {filename}")

# Abrir automáticamente (solo Windows) sino tienes windows comenta aqui porfis 
try:
    os.startfile(filename)
except Exception as e:
    print(f"⚠️ No se pudo abrir automáticamente: {e}")
