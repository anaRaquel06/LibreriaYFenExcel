"""
@author: Ana López
"""
#Datos de 2024 agrupados por acciones 
import yfinance as yf
from datetime import datetime
from openpyxl import Workbook #esta biblioteca pnecesita instalacion 
import os

# Lista de acciones del código que queremos extraer 
tickers = [
    "AA","AAL","AAPL","AMM.TO","ABBV","ABNB","ACTINVRB.MX","AC","AFRM",
    "AGNC","ALFAA.MX","ALPEKA.MX","ALSEA.MX","AMAT","AMD","AMX","AMZN",
    "APA","ASURB.MX","ATER","ATOS","AIY.DE","AVGO","AXP","BABA","BAC",
    "BA","BBAJIOO.MX","BIMBOA.MX","BMY","BNGO","CAT","CCL",
    "CEMEXCPO.MX","CHDRAUIB.MX","CLF","COST","CRM","CSCO",
    "CUERVO.MX","CVS","CVX","C","DAL","DIS","DVN","ELEKTRA.MX","ETSY",
    "FANG","FCX","FDX","FEMSAUBD.MX","FIBRAMQ12.MX","FIBRAPL14.MX",
    "FSLR","FUBO","FUNO11.MX","F","GAPB.MX","GCARSOA1.MX","GCC",
    "GENTERA.MX","GE","GFINBURO.MX","GFNORTEO.MX","GILD","GMEXICOB.MX",
    "GME","GM","GOLD","GOOGL","GRUMAB.MX","HD","INTC","JNJ","JPM",
    "KIMBERA.MX","KOFUBL.MX","KO","LABB.MX",
    "LASITEB-1.MX","LCID","LIVEPOLC-1.MX","LLY","LUV","LVS","LYFT","MARA",
    "MA","MCD","MEGACPO.MX","MELIN.MX","META","MFRISCOA-1.MX","MGM",
    "MRK","MRNA","MRO","MSFT","MU","NCLHN.MX","NFLX","NKE","NKLA","NUN.MX",
    "NVAX","NVDA","OMAB.MX","ORBIA.MX","ORCL","OXY1.MX","PARA","PBRN.MX","PE&OLES.MX",
    "PEP","PFE","PG","PINFRA.MX","PINS","PLTR","PYPL","QCOM","Q.MX","RCL",
    "RIOT","RIVN","ROKU","RA.MX","SBUX","SHOP","SITES1A-1.MX","SKLZ",
    "SOFI","SPCE","SQ","TALN.MX","TERRA13.MX","TGT","TLEVISACPO.MX","TMO",
    "TSLA","TSMN.MX","TWLO","TX","T","UAL","UBER","UNH","UPST","VESTA.MX",
    "VOLARA.MX","VZ","V","WALMEX.MX","WFC","WMT","WYNN","XOM","X","ZM"
]

# ------------------ AQUI CAMBIAS EL AÑO ------------------
year = 2024  # 👈 Cambia este valor para otro año
start_date = datetime(year, 1, 1)
end_date = datetime(year, 12, 31)
# ----------------------------------------------------------

# Crear libro de Excel
wb = Workbook()
ws = wb.active
ws.title = f"Resumen {year}"
headers = ["Ticker", "Fecha", "Apertura", "Alta", "Baja", "Cierre", "Volumen"]
ws.append(headers)

# Descargar datos de Yahoo Finance
for ticker in tickers:
    try:
        stock = yf.Ticker(ticker)
        df = stock.history(start=start_date, end=end_date)

        if df.empty:
            print(f"⚠️ Sin datos para {ticker}")
            continue

        df.reset_index(inplace=True)

        for _, row in df.iterrows():
            ws.append([
                ticker,
                row["Date"].strftime('%Y-%m-%d'),
                round(row["Open"], 2),
                round(row["High"], 2),
                round(row["Low"], 2),
                round(row["Close"], 2),
                int(row["Volume"])
            ])
        print(f"✅ {ticker} completado") ##otros indicadores (se ve amigable) 🟢👍💯
    except Exception as e:
        print(f"❌ Error con {ticker}: {e}") #otras opciones que podemos usar aqwui 🚨🚧⛔📢💣

# Ordenar datos: 1) Ticker, 2) Fecha ascendente
all_data = list(ws.iter_rows(min_row=2, values_only=True))
sorted_data = sorted(all_data, key=lambda row: (row[0], row[1]))

# Limpiar hoja y reescribir datos ordenados
ws.delete_rows(2, ws.max_row)
for row in sorted_data:
    ws.append(row)

# Guardar archivo
filename = f"historico_{year}.xlsx" #si es 2024 deberia ser historico_2024.xlsx se abrira solo 
wb.save(filename)

print(f"\n📁 Archivo guardado como: {filename}")

# Abrir automáticamente (solo Windows)
try:
    os.startfile(filename)
except Exception as e:
    print(f"⚠️ No se pudo abrir automáticamente: {e}")
