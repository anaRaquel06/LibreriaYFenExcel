"""
Created on Monday Aug  4 11:04:10 2025

@author: Ana Lopez 
"""
#El primer codigo guardaba los datos como la hoga de google que hicimos pero esto los guarda para en el historial del ultimo mes 
#este codigo servuira para las siguientes acciones 
import yfinance as yf
from datetime import datetime, timedelta
from openpyxl import Workbook
import os

# Lista de más de 140 acciones
tickers = [
    "AAPL", "MSFT", "GOOGL", "AMZN", "META", "TSLA", "NVDA", "BRK-B", "JPM", "JNJ",
    "V", "PG", "UNH", "MA", "HD", "DIS", "BAC", "XOM", "VZ", "ADBE",
    "NFLX", "KO", "T", "PFE", "PEP", "INTC", "CSCO", "MRK", "ABT", "CRM",
    "CMCSA", "CVX", "NKE", "WMT", "TMO", "MCD", "MDT", "ORCL", "AMGN", "COST",
    "QCOM", "HON", "DHR", "LLY", "ACN", "TXN", "LIN", "NEE", "PM", "IBM",
    "SBUX", "LOW", "BA", "AVGO", "RTX", "INTU", "AMAT", "ISRG", "GILD", "CAT",
    "LMT", "GE", "BKNG", "NOW", "DE", "SPGI", "MS", "BLK", "ADI", "MU",
    "ZTS", "SYK", "MDLZ", "ADP", "TGT", "USB", "MMM", "CI", "REGN", "CB",
    "CHTR", "VRTX", "CSX", "FIS", "GM", "F", "FDX", "ETN", "APD", "PLD",
    "BDX", "AON", "MO", "COF", "SO", "HCA", "ECL", "AEP", "ITW", "PSX",
    "PNC", "WMB", "D", "DOW", "CL", "EW", "C", "SHW", "ADM", "MNST",
    "WELL", "EL", "EXC", "ROST", "KHC", "VLO", "HPQ", "DAL", "GIS", "ALL",
    "KMB", "AZO", "YUM", "ILMN", "HES", "SLB", "SPG", "TT", "CNC", "BIIB"
]

# Fechas del último mes
end_date = datetime.today()
start_date = end_date - timedelta(days=30) #aqui le pedimos que nos traiga los ultimos 30 dias 

# Crear libro de Excel con Openpyxl
wb = Workbook()
ws = wb.active
ws.title = "Resumen Último Mes"
headers = ["Ticker", "Fecha", "Apertura", "Alta", "Baja", "Cierre", "Volumen"]
ws.append(headers)

# Descargar datos de yahoo finance con los tickers 
for ticker in tickers:
    try:
        stock = yf.Ticker(ticker)
        df = stock.history(start=start_date, end=end_date)

        if df.empty: #en lo personla se ve lindo en la temrinal usar estos emojis pero podemos cambialrlos 🚨🚧⛔📢💣
            print(f"⚠️ Sin datos para {ticker}")
            continue

        df.reset_index(inplace=True)

        for _, row in df.iterrows():
            ws.append([
                ticker,
                row["Date"].strftime('%Y-%m-%d'),
                round(row["Open"], 2),
                round(row["High"], 2),
                round(row["Low"], 2),
                round(row["Close"], 2),
                int(row["Volume"])
            ])
        print(f"✅ {ticker} completado") #otros indicadores (se ve amigable) 🟢👍💯
    except Exception as e:
        print(f"❌ Error con {ticker}: {e}") #otras opciones que podemos usar aqwui 🚨🚧⛔📢💣

# Ordenmaos por el dato cierre de mayor a menor 
# Tomamos todos los datos menos el encabezado 
all_data = list(ws.iter_rows(min_row=2, values_only=True))

# Ordenamos la columna cierre idx=5
#Nota tecnica= idc=indice 
sorted_data = sorted(all_data, key=lambda row: row[5], reverse=True)

# Se borran todas las filas menos el encabezado
ws.delete_rows(2, ws.max_row)

# Reescribir los datos ordenados ya que los quitamos 
for row in sorted_data:
    ws.append(row)
#otra nota, el metodo append, se usar para agregar, por lo que se si quiere agregar algo mas solo escribir:
#Nombredelahoja.apped(fila/columna) en ingles obvio 


# Guardar y abrir archivo
filename = "historico_ult_mes.xlsx"
wb.save(filename)

print(f"\n📁 Archivo guardado como: {filename}")


# Abrir automáticamente (solo en Windows)
#Aguregue esta ultimo funcion con la biblioteca os para no tener que buscarlo en las carptetas y solamente abrirlas 
try:
    os.startfile(filename)
except Exception as e:
    print(f"⚠️ No se pudo abrir el archivo automáticamente: {e}")



#POSIBLES ERRORES Y NOTAS:
#  a) Uno de los errores mas comunes a la hora de ejecutar es por que se tiene el libro en cuestion abierto 
#la biblioteca openpy no deja ejecutar si ese libro esta abierto y marcara el error: [Errno 13] Permission denied: 'historico_ult_mes.xlsx'
# b) La idea de este codigo es que este se ejecute cada cierto tiempo y usarlo para llevar a cabo los demas codigos     

